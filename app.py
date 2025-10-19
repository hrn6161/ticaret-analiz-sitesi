from flask import Flask, request, jsonify, render_template, send_file
import requests
from bs4 import BeautifulSoup
import time
import random
import re
from openpyxl import Workbook
from openpyxl.styles import Font
import sys
import logging
import os
from datetime import datetime
import cloudscraper
import urllib.parse

app = Flask(__name__)

print("🚀 AKILLI FİRMA ANALİZ SİSTEMİ BAŞLATILIYOR...")

# Logging setup
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class Config:
    def __init__(self):
        self.MAX_RESULTS = 8
        self.REQUEST_TIMEOUT = 30
        self.RETRY_ATTEMPTS = 2
        self.MAX_GTIP_CHECK = 3
        self.USER_AGENTS = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        ]

class SmartCrawler:
    def __init__(self, config):
        self.config = config
        self.scraper = cloudscraper.create_scraper()
    
    def smart_crawl(self, url, target_country):
        """Akıllı crawl - sadece ticari siteler"""
        logging.info(f"🌐 Crawl: {url[:60]}...")
        
        # Önce domain kontrolü - sadece ticari siteler
        if not self._is_commercial_domain(url):
            logging.info(f"🔍 Ticari olmayan site atlandı: {url}")
            return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': 'NON_COMMERCIAL'}
        
        time.sleep(1)
        
        result = self._try_cloudscraper(url, target_country)
        if result['status_code'] == 200:
            return result
        
        result = self._try_requests(url, target_country)
        if result['status_code'] == 200:
            return result
        
        logging.info(f"🔍 Sayfa erişilemiyor, snippet analizi kullanılıyor...")
        return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': 'BLOCKED'}
    
    def _is_commercial_domain(self, url):
        """Sadece ticari ve ticaret sitelerine izin ver"""
        commercial_domains = [
            'eximpedia', 'trademo', 'volza', 'exportgenius', 'comtrade',
            'alibaba', 'tradeindia', 'indiamart', 'go4worldbusiness',
            'kompass', 'globaltrade', 'worldtrade', 'tradekey',
            'companylist', 'exporters', 'suppliers', 'manufacturers',
            'business', 'trade', 'export', 'import', 'commerce',
            'customs', 'shipping', 'logistics', 'freight'
        ]
        
        spam_domains = [
            'donanımhaber', 'forum', 'blog', 'pdf', 'edu', 'academia',
            'wikipedia', 'social', 'facebook', 'twitter', 'instagram',
            'youtube', 'reddit', 'pinterest', 'tumblr'
        ]
        
        domain = url.lower()
        
        # Spam domain'leri engelle
        if any(spam in domain for spam in spam_domains):
            return False
        
        # Ticari domain'leri kabul et
        if any(commercial in domain for commercial in commercial_domains):
            return True
        
        return False
    
    def _try_cloudscraper(self, url, target_country):
        """Cloudscraper ile dene"""
        try:
            headers = {
                'User-Agent': random.choice(self.config.USER_AGENTS),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            }
            
            response = self.scraper.get(url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                logging.info(f"✅ Cloudscraper başarılı: {url}")
                return self._parse_content(response.text, target_country, response.status_code)
            else:
                return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': response.status_code}
                
        except Exception as e:
            logging.error(f"❌ Cloudscraper hatası: {e}")
            return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': 'ERROR'}
    
    def _try_requests(self, url, target_country):
        """Normal requests ile dene"""
        try:
            headers = {
                'User-Agent': random.choice(self.config.USER_AGENTS),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            }
            
            response = requests.get(url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                logging.info(f"✅ Requests başarılı: {url}")
                return self._parse_content(response.text, target_country, response.status_code)
            else:
                return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': response.status_code}
                
        except Exception as e:
            logging.error(f"❌ Requests hatası: {e}")
            return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': 'ERROR'}
    
    def _parse_content(self, html, target_country, status_code):
        """İçerik analizi - geliştirilmiş"""
        try:
            soup = BeautifulSoup(html, 'html.parser')
            text_content = soup.get_text()
            text_lower = text_content.lower()
            
            country_found = self._check_country(text_lower, target_country)
            gtip_codes = self.extract_gtip_codes(text_content)
            
            logging.info(f"🔍 Sayfa analizi: Ülke={country_found}, GTIP={gtip_codes[:3]}")
            
            return {
                'country_found': country_found,
                'gtip_codes': gtip_codes,
                'content_preview': text_content[:200] + "..." if len(text_content) > 200 else text_content,
                'status_code': status_code
            }
        except Exception as e:
            logging.error(f"❌ Parse hatası: {e}")
            return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': 'PARSE_ERROR'}
    
    def _check_country(self, text_lower, target_country):
        """Ülke kontrolü - daha spesifik"""
        country_variations = {
            'russia': ['russia', 'rusya', 'russian', 'rus', 'rossiya', 'rf'],
            'china': ['china', 'çin', 'chinese'],
            'iran': ['iran', 'irani', 'persian']
        }
        
        if target_country.lower() in country_variations:
            variations = country_variations[target_country.lower()]
            for variation in variations:
                if variation in text_lower:
                    return True
        
        return False
    
    def extract_gtip_codes(self, text):
        """GTIP kod çıkarma - daha doğru"""
        # Sadece gerçek GTIP formatları
        patterns = [
            r'\b\d{4}\.\d{2}\b',  # 8708.29 gibi
            r'\bGTIP[: ]*(\d{4}\.?\d{0,4})\b',
            r'\bHS[: ]*CODE[: ]*(\d{4}\.?\d{0,4})\b',
            r'\bH\.S\. Code[: ]*(\d{4}\.?\d{0,4})\b',
        ]
        
        all_codes = set()
        
        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                if isinstance(match, tuple):
                    match = match[0]
                
                code = re.sub(r'[^\d]', '', match)
                # Sadece 4-8 haneli ve anlamlı GTIP'ler
                if len(code) >= 4 and len(code) <= 8:
                    # Yılları filtrele (2023, 2024, 2025 gibi)
                    if not self._is_year(code):
                        all_codes.add(code[:4])  # İlk 4 hane
        
        return list(all_codes)
    
    def _is_year(self, code):
        """Yıl olup olmadığını kontrol et"""
        if len(code) == 4:
            year = int(code)
            # 1900-2030 arası yılları filtrele
            if 1900 <= year <= 2030:
                return True
        return False

class QualitySearcher:
    """Kaliteli arama - DuckDuckGo alternatifleri"""
    
    def __init__(self, config):
        self.config = config
        self.scraper = cloudscraper.create_scraper()
        logging.info("🔍 ÇOKLU ARAMA MOTORU HAZIR!")
    
    def search_quality(self, query, max_results=8):
        """Çoklu arama motoru desteği"""
        logging.info(f"🔍 Çoklu arama: {query}")
        
        time.sleep(2)
        
        # 1. DuckDuckGo ile dene
        results = self._search_duckduckgo_lite(query, max_results)
        if results:
            logging.info(f"✅ DuckDuckGo: {len(results)} sonuç")
            return results
        
        # 2. Yandex ile dene
        results = self._search_yandex(query, max_results)
        if results:
            logging.info(f"✅ Yandex: {len(results)} sonuç")
            return results
        
        # 3. Google benzeri arama
        results = self._search_generic(query, max_results)
        if results:
            logging.info(f"✅ Generic: {len(results)} sonuç")
            return results
        
        logging.info("❌ Arama sonucu bulunamadı")
        return []
    
    def _search_duckduckgo_lite(self, query, max_results):
        """DuckDuckGo Lite"""
        try:
            url = "https://lite.duckduckgo.com/lite/"
            data = {
                'q': query,
                'b': '',
            }
            
            headers = {
                'User-Agent': random.choice(self.config.USER_AGENTS),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Content-Type': 'application/x-www-form-urlencoded',
            }
            
            response = self.scraper.post(url, data=data, headers=headers, timeout=15)
            
            if response.status_code == 200:
                return self._parse_duckduckgo_results(response.text, max_results)
            else:
                logging.warning(f"❌ DuckDuckGo hatası {response.status_code}")
                return []
                
        except Exception as e:
            logging.error(f"❌ DuckDuckGo hatası: {e}")
            return []
    
    def _search_yandex(self, query, max_results):
        """Yandex arama"""
        try:
            url = "https://yandex.com/search/"
            params = {
                'text': f'{query} site:eximpedia.app OR site:trademo.com OR site:exportgenius.in'
            }
            
            headers = {
                'User-Agent': random.choice(self.config.USER_AGENTS),
            }
            
            response = self.scraper.get(url, params=params, headers=headers, timeout=15)
            
            if response.status_code == 200:
                return self._parse_yandex_results(response.text, max_results)
            else:
                logging.warning(f"❌ Yandex hatası {response.status_code}")
                return []
                
        except Exception as e:
            logging.error(f"❌ Yandex hatası: {e}")
            return []
    
    def _search_generic(self, query, max_results):
        """Genel arama - ticaret sitelerine özel"""
        try:
            # Ticaret sitelerini direkt ziyaret et
            trade_sites = [
                'https://eximpedia.app/search?q=',
                'https://trademo.com/search?query=',
                'https://www.exportgenius.in/search?q='
            ]
            
            all_results = []
            
            for site in trade_sites[:2]:  # İlk 2 siteyi dene
                try:
                    search_url = site + urllib.parse.quote(query)
                    headers = {
                        'User-Agent': random.choice(self.config.USER_AGENTS),
                    }
                    
                    response = self.scraper.get(search_url, headers=headers, timeout=10)
                    
                    if response.status_code == 200:
                        site_results = self._parse_trade_site_results(response.text, search_url, max_results)
                        all_results.extend(site_results)
                        
                        if len(all_results) >= max_results:
                            break
                            
                except Exception as e:
                    logging.warning(f"❌ Site arama hatası {site}: {e}")
                    continue
            
            return all_results[:max_results]
                
        except Exception as e:
            logging.error(f"❌ Generic arama hatası: {e}")
            return []
    
    def _parse_duckduckgo_results(self, html, max_results):
        """DuckDuckGo sonuçlarını parse et"""
        soup = BeautifulSoup(html, 'html.parser')
        results = []
        
        rows = soup.find_all('tr')
        
        for i in range(0, len(rows)-1, 3):
            if len(results) >= max_results:
                break
                
            try:
                title_row = rows[i]
                link_elem = title_row.find('a', href=True)
                if not link_elem:
                    continue
                    
                title = link_elem.get_text(strip=True)
                url = link_elem.get('href')
                
                if not url or not self._is_quality_url(url):
                    continue
                
                snippet_row = rows[i+1] if i+1 < len(rows) else None
                snippet = ""
                if snippet_row:
                    snippet_elem = snippet_row.find('td')
                    if snippet_elem:
                        snippet = snippet_elem.get_text(strip=True)
                
                results.append({
                    'title': title,
                    'url': url,
                    'snippet': snippet,
                    'full_text': f"{title} {snippet}",
                    'domain': self._extract_domain(url),
                    'search_engine': 'duckduckgo'
                })
                
            except Exception as e:
                continue
        
        return results
    
    def _parse_yandex_results(self, html, max_results):
        """Yandex sonuçlarını parse et"""
        soup = BeautifulSoup(html, 'html.parser')
        results = []
        
        # Yandex result container
        result_blocks = soup.find_all('li', class_='serp-item') or soup.find_all('div', class_='organic')
        
        for block in result_blocks:
            if len(results) >= max_results:
                break
                
            try:
                link_elem = block.find('a', href=True)
                if not link_elem:
                    continue
                    
                title = link_elem.get_text(strip=True)
                url = link_elem.get('href')
                
                if not url or not self._is_quality_url(url):
                    continue
                
                snippet_elem = block.find('div', class_='text-container') or block.find('div', class_='organic__text')
                snippet = snippet_elem.get_text(strip=True) if snippet_elem else ""
                
                results.append({
                    'title': title,
                    'url': url,
                    'snippet': snippet,
                    'full_text': f"{title} {snippet}",
                    'domain': self._extract_domain(url),
                    'search_engine': 'yandex'
                })
                
            except Exception as e:
                continue
        
        return results
    
    def _parse_trade_site_results(self, html, search_url, max_results):
        """Ticaret sitesi sonuçlarını parse et"""
        soup = BeautifulSoup(html, 'html.parser')
        results = []
        
        # Basit link extraction
        links = soup.find_all('a', href=True)
        
        for link in links:
            if len(results) >= max_results:
                break
                
            try:
                title = link.get_text(strip=True)
                url = link.get('href')
                
                if not title or len(title) < 10:
                    continue
                    
                # URL'yi tamamla
                if url.startswith('/'):
                    from urllib.parse import urlparse
                    base_url = urlparse(search_url).scheme + '://' + urlparse(search_url).netloc
                    url = base_url + url
                
                if not self._is_quality_url(url):
                    continue
                
                results.append({
                    'title': title,
                    'url': url,
                    'snippet': f"Trade site result for {title}",
                    'full_text': title,
                    'domain': self._extract_domain(url),
                    'search_engine': 'trade_site'
                })
                
            except Exception as e:
                continue
        
        return results
    
    def _is_quality_url(self, url):
        """Kaliteli URL kontrolü"""
        quality_domains = [
            'eximpedia.app', 'trademo.com', 'volza.com', 'exportgenius.in',
            'comtrade.un.org', 'alibaba.com', 'tradeindia.com',
            'kompass.com', 'go4worldbusiness.com'
        ]
        
        domain = self._extract_domain(url)
        return any(quality_domain in domain for quality_domain in quality_domains)
    
    def _extract_domain(self, url):
        """Domain çıkar"""
        try:
            from urllib.parse import urlparse
            return urlparse(url).netloc
        except:
            return ""

class ExactQueryGenerator:
    """TAM FİRMA ADLI sorgu generator"""
    
    @staticmethod
    def generate_queries(company, country):
        """TAM FİRMA ADI ile kaliteli sorgular"""
        
        queries = [
            f'"{company}" {country} export',
            f'"{company}" {country} import',
            f'"{company}" {country}',
            f'"{company}" trade',
            f'"{company}" customs',
            f'{company} {country} supplier',
            f'{company} {country} manufacturer'
        ]
        
        logging.info(f"🔍 {len(queries)} sorgu oluşturuldu")
        return queries

class QuickEURLexChecker:
    def __init__(self, config):
        self.config = config
        self.sanction_cache = {}
    
    def quick_check_gtip(self, gtip_codes):
        """GTIP kontrolü"""
        if not gtip_codes:
            return []
            
        sanctioned_codes = []
        checked_codes = gtip_codes[:self.config.MAX_GTIP_CHECK]
        
        logging.info(f"🔍 EUR-Lex kontrolü: {checked_codes}")
        
        for gtip_code in checked_codes:
            if gtip_code in self.sanction_cache:
                if self.sanction_cache[gtip_code]:
                    sanctioned_codes.append(gtip_code)
                continue
                
            try:
                time.sleep(1)
                
                url = "https://eur-lex.europa.eu/search.html"
                params = {
                    'text': f'"{gtip_code}" Russia sanction',
                    'type': 'advanced',
                    'lang': 'en'
                }
                
                headers = {
                    'User-Agent': random.choice(self.config.USER_AGENTS),
                }
                
                response = requests.get(url, params=params, headers=headers, timeout=10)
                
                if response.status_code == 200:
                    soup = BeautifulSoup(response.text, 'html.parser')
                    content = soup.get_text().lower()
                    
                    sanction_terms = ['prohibited', 'banned', 'sanction', 'restricted', 'embargo']
                    found_sanction = any(term in content for term in sanction_terms)
                    
                    if found_sanction:
                        sanctioned_codes.append(gtip_code)
                        self.sanction_cache[gtip_code] = True
                        logging.info(f"⛔ Yaptırımlı kod: {gtip_code}")
                    else:
                        self.sanction_cache[gtip_code] = False
                
            except Exception as e:
                logging.error(f"❌ EUR-Lex kontrol hatası: {e}")
                continue
        
        return sanctioned_codes

class SmartTradeAnalyzer:
    def __init__(self, config):
        self.config = config
        self.searcher = QualitySearcher(config)
        self.crawler = SmartCrawler(config)
        self.eur_lex_checker = QuickEURLexChecker(config)
        self.query_generator = ExactQueryGenerator()
    
    def smart_analyze(self, company, country):
        """AKILLI ANALİZ"""
        logging.info(f"🤖 AKILLI ANALİZ: '{company}' ↔ {country}")
        
        search_queries = self.query_generator.generate_queries(company, country)
        
        all_results = []
        found_urls = set()
        country_connection_found = False
        
        for i, query in enumerate(search_queries, 1):
            try:
                logging.info(f"🔍 Sorgu {i}/{len(search_queries)}: {query}")
                
                if i > 1:
                    time.sleep(2)
                
                search_results = self.searcher.search_quality(query, self.config.MAX_RESULTS)
                
                if not search_results:
                    logging.warning(f"⚠️ Sonuç bulunamadı: {query}")
                    continue
                
                for j, result in enumerate(search_results, 1):
                    if result['url'] in found_urls:
                        continue
                    
                    found_urls.add(result['url'])
                    
                    logging.info(f"📄 Sonuç {j}: {result['title'][:50]}...")
                    
                    if j > 1:
                        time.sleep(1)
                    
                    crawl_result = self.crawler.smart_crawl(result['url'], country)
                    
                    if crawl_result['status_code'] == 'NON_COMMERCIAL':
                        continue
                    
                    if crawl_result['country_found']:
                        country_connection_found = True
                        logging.info(f"🚨 ÜLKE BAĞLANTISI: {company} ↔ {country}")
                    
                    sanctioned_gtips = []
                    if crawl_result['gtip_codes']:
                        sanctioned_gtips = self.eur_lex_checker.quick_check_gtip(crawl_result['gtip_codes'])
                    
                    confidence = self._calculate_confidence(crawl_result, sanctioned_gtips, result['domain'])
                    
                    analysis = self.create_analysis_result(
                        company, country, result, crawl_result, sanctioned_gtips, confidence, country_connection_found
                    )
                    
                    all_results.append(analysis)
                    
                    if len(all_results) >= 3:
                        logging.info("🎯 3 sonuç bulundu")
                        return all_results
                
            except Exception as e:
                logging.error(f"❌ Sorgu hatası: {e}")
                continue
        
        return all_results
    
    def _calculate_confidence(self, crawl_result, sanctioned_gtips, domain):
        """Güven seviyesi"""
        confidence = 0
        
        if crawl_result['country_found']:
            confidence += 50
        
        if crawl_result['gtip_codes']:
            confidence += 30
        
        if sanctioned_gtips:
            confidence += 20
        
        return min(confidence, 100)
    
    def create_analysis_result(self, company, country, search_result, crawl_result, sanctioned_gtips, confidence, country_connection_found):
        """Analiz sonucu"""
        
        if country_connection_found:
            if sanctioned_gtips:
                status = "YÜKSEK_RİSK"
                explanation = f"⛔ YÜKSEK RİSK: {company} şirketi {country} ile yasaklı ürün ticareti yapıyor"
                ai_tavsiye = f"🔴 ACİL İNCELEME! Yasaklı GTIP: {', '.join(sanctioned_gtips)}"
                risk_level = "YÜKSEK"
            else:
                status = "RİSK_VAR"
                explanation = f"🟡 RİSK VAR: {company} şirketi {country} ile ticaret bağlantısı var"
                ai_tavsiye = "Ticaret bağlantısı doğrulandı. Detaylı inceleme önerilir."
                risk_level = "ORTA"
        else:
            status = "TEMİZ"
            explanation = f"✅ TEMİZ: {company} şirketinin {country} ile bağlantısı bulunamadı"
            ai_tavsiye = "Risk bulunamadı"
            risk_level = "YOK"
        
        return {
            'ŞİRKET': company,
            'ÜLKE': country,
            'DURUM': status,
            'AI_AÇIKLAMA': explanation,
            'AI_TAVSIYE': ai_tavsiye,
            'YAPTIRIM_RISKI': risk_level,
            'TESPIT_EDILEN_GTIPLER': ', '.join(crawl_result['gtip_codes'][:3]),
            'YAPTIRIMLI_GTIPLER': ', '.join(sanctioned_gtips),
            'ULKE_BAGLANTISI': 'EVET' if crawl_result['country_found'] else 'HAYIR',
            'BAŞLIK': search_result['title'],
            'URL': search_result['url'],
            'ÖZET': search_result['snippet'],
            'GÜVEN_SEVİYESİ': f"%{confidence}",
            'KAYNAK_TİPİ': search_result.get('search_engine', 'UNKNOWN')
        }

def create_excel_report(results, company, country):
    """Excel raporu"""
    try:
        filename = f"{company.replace(' ', '_')}_{country}_analiz.xlsx"
        filepath = os.path.join('/tmp', filename)
        
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Analiz Sonuçları"
        
        headers = [
            'ŞİRKET', 'ÜLKE', 'DURUM', 'YAPTIRIM_RISKI', 'ULKE_BAGLANTISI',
            'TESPIT_EDILEN_GTIPLER', 'YAPTIRIMLI_GTIPLER', 'GÜVEN_SEVİYESİ',
            'AI_AÇIKLAMA', 'AI_TAVSIYE', 'BAŞLIK', 'URL'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row, result in enumerate(results, 2):
            ws1.cell(row=row, column=1, value=str(result.get('ŞİRKET', '')))
            ws1.cell(row=row, column=2, value=str(result.get('ÜLKE', '')))
            ws1.cell(row=row, column=3, value=str(result.get('DURUM', '')))
            ws1.cell(row=row, column=4, value=str(result.get('YAPTIRIM_RISKI', '')))
            ws1.cell(row=row, column=5, value=str(result.get('ULKE_BAGLANTISI', '')))
            ws1.cell(row=row, column=6, value=str(result.get('TESPIT_EDILEN_GTIPLER', '')))
            ws1.cell(row=row, column=7, value=str(result.get('YAPTIRIMLI_GTIPLER', '')))
            ws1.cell(row=row, column=8, value=str(result.get('GÜVEN_SEVİYESİ', '')))
            ws1.cell(row=row, column=9, value=str(result.get('AI_AÇIKLAMA', '')))
            ws1.cell(row=row, column=10, value=str(result.get('AI_TAVSIYE', '')))
            ws1.cell(row=row, column=11, value=str(result.get('BAŞLIK', '')))
            ws1.cell(row=row, column=12, value=str(result.get('URL', '')))
        
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        logging.info(f"✅ Excel raporu oluşturuldu: {filepath}")
        return filepath
        
    except Exception as e:
        logging.error(f"❌ Excel rapor hatası: {e}")
        return None

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        start_time = time.time()
        
        data = request.get_json()
        company = data.get('company', '').strip()
        country = data.get('country', '').strip()
        
        if not company or not country:
            return jsonify({"error": "Şirket ve ülke bilgisi gereklidir"}), 400
        
        logging.info(f"🚀 ANALİZ BAŞLATILIYOR: '{company}' - {country}")
        
        config = Config()
        analyzer = SmartTradeAnalyzer(config)
        
        results = analyzer.smart_analyze(company, country)
        
        excel_filepath = create_excel_report(results, company, country)
        
        execution_time = time.time() - start_time
        
        response_data = {
            "success": True,
            "company": company,
            "country": country,
            "execution_time": f"{execution_time:.2f}s",
            "total_results": len(results),
            "analysis": results,
            "excel_download_url": f"/download-excel?company={company}&country={country}" if excel_filepath else None
        }
        
        return jsonify(response_data)
        
    except Exception as e:
        logging.error(f"❌ Analiz hatası: {e}")
        return jsonify({"error": f"Sunucu hatası: {str(e)}"}), 500

@app.route('/download-excel')
def download_excel():
    try:
        company = request.args.get('company', '')
        country = request.args.get('country', '')
        
        filename = f"{company.replace(' ', '_')}_{country}_analiz.xlsx"
        filepath = os.path.join('/tmp', filename)
        
        if os.path.exists(filepath):
            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify({"error": "Excel dosyası bulunamadı"}), 404
            
    except Exception as e:
        logging.error(f"❌ Excel indirme hatası: {e}")
        return jsonify({"error": f"İndirme hatası: {str(e)}"}), 500

@app.route('/health')
def health():
    return jsonify({"status": "healthy", "timestamp": datetime.now().isoformat()})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
