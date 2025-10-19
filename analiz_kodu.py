import requests
from bs4 import BeautifulSoup
import time
import random
import re
from openpyxl import Workbook
from openpyxl.styles import Font
import sys
import logging
import os
from datetime import datetime
import cloudscraper
import urllib.parse

print("🚀 GERÇEK TİCARET VERİTABANI ANALİZ SİSTEMİ")

# Logging setup
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
        logging.StreamHandler()
    ]
)

class Config:
    def __init__(self):
        self.MAX_RESULTS = 8
        self.REQUEST_TIMEOUT = 30
        self.RETRY_ATTEMPTS = 2
        self.MAX_GTIP_CHECK = 3
        self.USER_AGENTS = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        ]

class RealTradeDatabaseSearcher:
    """GERÇEK TİCARET VERİTABANLARI"""
    
    def __init__(self, config):
        self.config = config
        self.scraper = cloudscraper.create_scraper()
        print("   🔍 GERÇEK TİCARET VERİTABANI BAĞLANTISI HAZIR!")
    
    def search_real_trade_data(self, company, country, max_results=8):
        """Gerçek ticaret veritabanlarında ara"""
        print(f"   🔍 Gerçek veritabanı arama: '{company}' ↔ {country}")
        
        all_results = []
        
        # 1. Eximpedia.app
        eximpedia_results = self._search_eximpedia(company, country, max_results)
        if eximpedia_results:
            all_results.extend(eximpedia_results)
            print(f"   ✅ Eximpedia: {len(eximpedia_results)} sonuç")
        
        # 2. Trademo.com
        trademo_results = self._search_trademo(company, country, max_results)
        if trademo_results:
            all_results.extend(trademo_results)
            print(f"   ✅ Trademo: {len(trademo_results)} sonuç")
        
        # 3. ExportGenius
        exportgenius_results = self._search_exportgenius(company, country, max_results)
        if exportgenius_results:
            all_results.extend(exportgenius_results)
            print(f"   ✅ ExportGenius: {len(exportgenius_results)} sonuç")
        
        # 4. Kompass
        kompass_results = self._search_kompass(company, country, max_results)
        if kompass_results:
            all_results.extend(kompass_results)
            print(f"   ✅ Kompass: {len(kompass_results)} sonuç")
        
        return all_results[:max_results]
    
    def _search_eximpedia(self, company, country, max_results):
        """Eximpedia.app"""
        try:
            search_url = f"https://eximpedia.app/search?q={urllib.parse.quote(company + ' ' + country)}"
            
            headers = {
                'User-Agent': random.choice(self.config.USER_AGENTS),
            }
            
            response = self.scraper.get(search_url, headers=headers, timeout=15)
            
            if response.status_code == 200:
                return self._parse_eximpedia_results(response.text, max_results)
            else:
                print(f"   ❌ Eximpedia erişilemedi: {response.status_code}")
                return []
                
        except Exception as e:
            print(f"   ❌ Eximpedia hatası: {e}")
            return []
    
    def _parse_eximpedia_results(self, html, max_results):
        """Eximpedia sonuçları"""
        soup = BeautifulSoup(html, 'html.parser')
        results = []
        
        result_cards = soup.find_all('div', class_=lambda x: x and ('card' in x or 'result' in x or 'item' in x))
        
        for card in result_cards[:max_results]:
            try:
                title_elem = card.find(['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'a'])
                if not title_elem:
                    continue
                
                title = title_elem.get_text(strip=True)
                if len(title) < 5:
                    continue
                
                link_elem = card.find('a', href=True)
                url = link_elem.get('href') if link_elem else ""
                if url and url.startswith('/'):
                    url = 'https://eximpedia.app' + url
                
                snippet_elem = card.find(['p', 'div', 'span'], class_=lambda x: x and ('desc' in x or 'text' in x))
                snippet = snippet_elem.get_text(strip=True) if snippet_elem else ""
                
                results.append({
                    'title': title,
                    'url': url,
                    'snippet': snippet,
                    'full_text': f"{title} {snippet}",
                    'domain': 'eximpedia.app',
                    'search_engine': 'eximpedia'
                })
                
            except Exception as e:
                continue
        
        if not results:
            links = soup.find_all('a', href=True)
            for link in links[:max_results]:
                try:
                    title = link.get_text(strip=True)
                    if len(title) < 10:
                        continue
                    
                    url = link.get('href')
                    if url.startswith('/'):
                        url = 'https://eximpedia.app' + url
                    
                    results.append({
                        'title': title,
                        'url': url,
                        'snippet': "Eximpedia trade data",
                        'full_text': title,
                        'domain': 'eximpedia.app',
                        'search_engine': 'eximpedia'
                    })
                except:
                    continue
        
        return results
    
    def _search_trademo(self, company, country, max_results):
        """Trademo.com"""
        try:
            search_url = f"https://trademo.com/search?query={urllib.parse.quote(company + ' ' + country)}"
            
            headers = {
                'User-Agent': random.choice(self.config.USER_AGENTS),
            }
            
            response = self.scraper.get(search_url, headers=headers, timeout=15)
            
            if response.status_code == 200:
                return self._parse_trademo_results(response.text, max_results)
            else:
                print(f"   ❌ Trademo erişilemedi: {response.status_code}")
                return []
                
        except Exception as e:
            print(f"   ❌ Trademo hatası: {e}")
            return []
    
    def _parse_trademo_results(self, html, max_results):
        """Trademo sonuçları"""
        soup = BeautifulSoup(html, 'html.parser')
        results = []
        
        result_items = soup.find_all('div', class_=lambda x: x and ('product' in x or 'item' in x or 'result' in x))
        
        for item in result_items[:max_results]:
            try:
                title_elem = item.find(['h1', 'h2', 'h3', 'h4', 'a'])
                if not title_elem:
                    continue
                
                title = title_elem.get_text(strip=True)
                if len(title) < 5:
                    continue
                
                link_elem = item.find('a', href=True)
                url = link_elem.get('href') if link_elem else ""
                if url and url.startswith('/'):
                    url = 'https://trademo.com' + url
                
                snippet_elem = item.find(['p', 'div'], class_=lambda x: x and ('description' in x or 'desc' in x))
                snippet = snippet_elem.get_text(strip=True) if snippet_elem else ""
                
                results.append({
                    'title': title,
                    'url': url,
                    'snippet': snippet,
                    'full_text': f"{title} {snippet}",
                    'domain': 'trademo.com',
                    'search_engine': 'trademo'
                })
                
            except Exception as e:
                continue
        
        return results
    
    def _search_exportgenius(self, company, country, max_results):
        """ExportGenius"""
        try:
            search_url = f"https://www.exportgenius.in/search?q={urllib.parse.quote(company + ' ' + country)}"
            
            headers = {
                'User-Agent': random.choice(self.config.USER_AGENTS),
            }
            
            response = self.scraper.get(search_url, headers=headers, timeout=15)
            
            if response.status_code == 200:
                return self._parse_exportgenius_results(response.text, max_results)
            else:
                print(f"   ❌ ExportGenius erişilemedi: {response.status_code}")
                return []
                
        except Exception as e:
            print(f"   ❌ ExportGenius hatası: {e}")
            return []
    
    def _parse_exportgenius_results(self, html, max_results):
        """ExportGenius sonuçları"""
        soup = BeautifulSoup(html, 'html.parser')
        results = []
        
        links = soup.find_all('a', href=True)
        for link in links[:max_results]:
            try:
                title = link.get_text(strip=True)
                if len(title) < 10:
                    continue
                
                url = link.get('href')
                if url and url.startswith('/'):
                    url = 'https://www.exportgenius.in' + url
                
                if 'exportgenius' not in url:
                    continue
                
                results.append({
                    'title': title,
                    'url': url,
                    'snippet': "ExportGenius trade data",
                    'full_text': title,
                    'domain': 'exportgenius.in',
                    'search_engine': 'exportgenius'
                })
                
            except Exception as e:
                continue
        
        return results
    
    def _search_kompass(self, company, country, max_results):
        """Kompass"""
        try:
            search_url = f"https://www.kompass.com/search/?text={urllib.parse.quote(company + ' ' + country)}"
            
            headers = {
                'User-Agent': random.choice(self.config.USER_AGENTS),
            }
            
            response = self.scraper.get(search_url, headers=headers, timeout=15)
            
            if response.status_code == 200:
                return self._parse_kompass_results(response.text, max_results)
            else:
                print(f"   ❌ Kompass erişilemedi: {response.status_code}")
                return []
                
        except Exception as e:
            print(f"   ❌ Kompass hatası: {e}")
            return []
    
    def _parse_kompass_results(self, html, max_results):
        """Kompass sonuçları"""
        soup = BeautifulSoup(html, 'html.parser')
        results = []
        
        links = soup.find_all('a', href=True)
        for link in links[:max_results]:
            try:
                title = link.get_text(strip=True)
                if len(title) < 10 or 'kompass' not in title.lower():
                    continue
                
                url = link.get('href')
                if url and url.startswith('/'):
                    url = 'https://www.kompass.com' + url
                
                results.append({
                    'title': title,
                    'url': url,
                    'snippet': "Kompass company directory",
                    'full_text': title,
                    'domain': 'kompass.com',
                    'search_engine': 'kompass'
                })
                
            except Exception as e:
                continue
        
        return results

class SmartCrawler:
    def __init__(self, config):
        self.config = config
        self.scraper = cloudscraper.create_scraper()
    
    def smart_crawl(self, url, target_country):
        """Akıllı crawl"""
        print(f"   🌐 Crawl: {url[:60]}...")
        
        time.sleep(1)
        
        result = self._try_cloudscraper(url, target_country)
        if result['status_code'] == 200:
            return result
        
        result = self._try_requests(url, target_country)
        if result['status_code'] == 200:
            return result
        
        return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': 'BLOCKED'}
    
    def _try_cloudscraper(self, url, target_country):
        """Cloudscraper ile dene"""
        try:
            headers = {
                'User-Agent': random.choice(self.config.USER_AGENTS),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            }
            
            response = self.scraper.get(url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                print(f"   ✅ Cloudscraper başarılı: {url}")
                return self._parse_content(response.text, target_country, response.status_code)
            else:
                return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': response.status_code}
                
        except Exception as e:
            print(f"   ❌ Cloudscraper hatası: {e}")
            return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': 'ERROR'}
    
    def _try_requests(self, url, target_country):
        """Normal requests ile dene"""
        try:
            headers = {
                'User-Agent': random.choice(self.config.USER_AGENTS),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            }
            
            response = requests.get(url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                print(f"   ✅ Requests başarılı: {url}")
                return self._parse_content(response.text, target_country, response.status_code)
            else:
                return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': response.status_code}
                
        except Exception as e:
            print(f"   ❌ Requests hatası: {e}")
            return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': 'ERROR'}
    
    def _parse_content(self, html, target_country, status_code):
        """İçerik analizi"""
        try:
            soup = BeautifulSoup(html, 'html.parser')
            text_content = soup.get_text()
            text_lower = text_content.lower()
            
            country_found = self._check_country(text_lower, target_country)
            gtip_codes = self.extract_gtip_codes(text_content)
            
            print(f"   🔍 Sayfa analizi: Ülke={country_found}, GTIP={gtip_codes[:3]}")
            
            return {
                'country_found': country_found,
                'gtip_codes': gtip_codes,
                'content_preview': text_content[:200] + "..." if len(text_content) > 200 else text_content,
                'status_code': status_code
            }
        except Exception as e:
            print(f"   ❌ Parse hatası: {e}")
            return {'country_found': False, 'gtip_codes': [], 'content_preview': '', 'status_code': 'PARSE_ERROR'}
    
    def _check_country(self, text_lower, target_country):
        """Ülke kontrolü"""
        country_variations = {
            'russia': ['russia', 'rusya', 'russian', 'rus', 'rossiya', 'rf'],
            'china': ['china', 'çin', 'chinese'],
            'iran': ['iran', 'irani', 'persian']
        }
        
        if target_country.lower() in country_variations:
            variations = country_variations[target_country.lower()]
            for variation in variations:
                if variation in text_lower:
                    return True
        
        return False
    
    def extract_gtip_codes(self, text):
        """GTIP kod çıkarma"""
        patterns = [
            r'\b\d{4}\.\d{2}\b',
            r'\bGTIP[: ]*(\d{4}\.?\d{0,4})\b',
            r'\bHS[: ]*CODE[: ]*(\d{4}\.?\d{0,4})\b',
        ]
        
        all_codes = set()
        
        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                if isinstance(match, tuple):
                    match = match[0]
                
                code = re.sub(r'[^\d]', '', match)
                if len(code) >= 4 and len(code) <= 8:
                    if not self._is_year(code):
                        all_codes.add(code[:4])
        
        return list(all_codes)
    
    def _is_year(self, code):
        """Yıl kontrolü"""
        if len(code) == 4:
            year = int(code)
            if 1900 <= year <= 2030:
                return True
        return False

class QuickEURLexChecker:
    def __init__(self, config):
        self.config = config
        self.sanction_cache = {}
    
    def quick_check_gtip(self, gtip_codes):
        """GTIP kontrolü"""
        if not gtip_codes:
            return []
            
        sanctioned_codes = []
        checked_codes = gtip_codes[:self.config.MAX_GTIP_CHECK]
        
        print(f"   🔍 EUR-Lex kontrolü: {checked_codes}")
        
        for gtip_code in checked_codes:
            if gtip_code in self.sanction_cache:
                if self.sanction_cache[gtip_code]:
                    sanctioned_codes.append(gtip_code)
                continue
                
            try:
                time.sleep(1)
                
                url = "https://eur-lex.europa.eu/search.html"
                params = {
                    'text': f'"{gtip_code}" Russia sanction',
                    'type': 'advanced',
                    'lang': 'en'
                }
                
                headers = {
                    'User-Agent': random.choice(self.config.USER_AGENTS),
                }
                
                response = requests.get(url, params=params, headers=headers, timeout=10)
                
                if response.status_code == 200:
                    soup = BeautifulSoup(response.text, 'html.parser')
                    content = soup.get_text().lower()
                    
                    sanction_terms = ['prohibited', 'banned', 'sanction', 'restricted', 'embargo']
                    found_sanction = any(term in content for term in sanction_terms)
                    
                    if found_sanction:
                        sanctioned_codes.append(gtip_code)
                        self.sanction_cache[gtip_code] = True
                        print(f"   ⛔ Yaptırımlı kod: {gtip_code}")
                    else:
                        self.sanction_cache[gtip_code] = False
                
            except Exception as e:
                print(f"   ❌ EUR-Lex kontrol hatası: {e}")
                continue
        
        return sanctioned_codes

class SmartTradeAnalyzer:
    def __init__(self, config):
        self.config = config
        self.searcher = RealTradeDatabaseSearcher(config)
        self.crawler = SmartCrawler(config)
        self.eur_lex_checker = QuickEURLexChecker(config)
    
    def smart_analyze(self, company, country):
        """GERÇEK VERİTABANI ANALİZİ"""
        print(f"🤖 GERÇEK VERİTABANI ANALİZİ: '{company}' ↔ {country}")
        
        search_results = self.searcher.search_real_trade_data(company, country, self.config.MAX_RESULTS)
        
        if not search_results:
            print("   ❌ Gerçek ticaret verisi bulunamadı")
            return []
        
        all_results = []
        country_connection_found = False
        
        for i, result in enumerate(search_results, 1):
            try:
                print(f"   📄 Veritabanı sonuç {i}: {result['title'][:50]}...")
                
                if i > 1:
                    time.sleep(1)
                
                crawl_result = self.crawler.smart_crawl(result['url'], country)
                
                if crawl_result['country_found']:
                    country_connection_found = True
                    print(f"   🚨 ÜLKE BAĞLANTISI: {company} ↔ {country}")
                
                sanctioned_gtips = []
                if crawl_result['gtip_codes']:
                    sanctioned_gtips = self.eur_lex_checker.quick_check_gtip(crawl_result['gtip_codes'])
                
                confidence = self._calculate_confidence(crawl_result, sanctioned_gtips, result['domain'])
                
                analysis = self.create_analysis_result(
                    company, country, result, crawl_result, sanctioned_gtips, confidence, country_connection_found
                )
                
                all_results.append(analysis)
                
                if len(all_results) >= 3:
                    print("   🎯 3 sonuç bulundu")
                    return all_results
                
            except Exception as e:
                print(f"   ❌ Sonuç işleme hatası: {e}")
                continue
        
        return all_results
    
    def _calculate_confidence(self, crawl_result, sanctioned_gtips, domain):
        """Güven seviyesi"""
        confidence = 0
        
        if crawl_result['country_found']:
            confidence += 50
        
        if crawl_result['gtip_codes']:
            confidence += 30
        
        if sanctioned_gtips:
            confidence += 20
        
        return min(confidence, 100)
    
    def create_analysis_result(self, company, country, search_result, crawl_result, sanctioned_gtips, confidence, country_connection_found):
        """Analiz sonucu"""
        
        if country_connection_found:
            if sanctioned_gtips:
                status = "YÜKSEK_RİSK"
                explanation = f"⛔ YÜKSEK RİSK: {company} şirketi {country} ile yasaklı ürün ticareti yapıyor"
                ai_tavsiye = f"🔴 ACİL İNCELEME! Yasaklı GTIP: {', '.join(sanctioned_gtips)}"
                risk_level = "YÜKSEK"
            else:
                status = "RİSK_VAR"
                explanation = f"🟡 RİSK VAR: {company} şirketi {country} ile ticaret bağlantısı var"
                ai_tavsiye = "Ticaret bağlantısı doğrulandı. Detaylı inceleme önerilir."
                risk_level = "ORTA"
        else:
            status = "TEMİZ"
            explanation = f"✅ TEMİZ: {company} şirketinin {country} ile bağlantısı bulunamadı"
            ai_tavsiye = "Risk bulunamadı"
            risk_level = "YOK"
        
        return {
            'ŞİRKET': company,
            'ÜLKE': country,
            'DURUM': status,
            'AI_AÇIKLAMA': explanation,
            'AI_TAVSIYE': ai_tavsiye,
            'YAPTIRIM_RISKI': risk_level,
            'TESPIT_EDILEN_GTIPLER': ', '.join(crawl_result['gtip_codes'][:3]),
            'YAPTIRIMLI_GTIPLER': ', '.join(sanctioned_gtips),
            'ULKE_BAGLANTISI': 'EVET' if crawl_result['country_found'] else 'HAYIR',
            'BAŞLIK': search_result['title'],
            'URL': search_result['url'],
            'ÖZET': search_result['snippet'],
            'GÜVEN_SEVİYESİ': f"%{confidence}",
            'KAYNAK_TİPİ': search_result.get('search_engine', 'REAL_TRADE_DB')
        }

def create_excel_report(results, company, country):
    """Excel raporu"""
    try:
        filename = f"{company.replace(' ', '_')}_{country}_analiz.xlsx"
        
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Analiz Sonuçları"
        
        headers = [
            'ŞİRKET', 'ÜLKE', 'DURUM', 'YAPTIRIM_RISKI', 'ULKE_BAGLANTISI',
            'TESPIT_EDILEN_GTIPLER', 'YAPTIRIMLI_GTIPLER', 'GÜVEN_SEVİYESİ',
            'AI_AÇIKLAMA', 'AI_TAVSIYE', 'BAŞLIK', 'URL'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row, result in enumerate(results, 2):
            ws1.cell(row=row, column=1, value=str(result.get('ŞİRKET', '')))
            ws1.cell(row=row, column=2, value=str(result.get('ÜLKE', '')))
            ws1.cell(row=row, column=3, value=str(result.get('DURUM', '')))
            ws1.cell(row=row, column=4, value=str(result.get('YAPTIRIM_RISKI', '')))
            ws1.cell(row=row, column=5, value=str(result.get('ULKE_BAGLANTISI', '')))
            ws1.cell(row=row, column=6, value=str(result.get('TESPIT_EDILEN_GTIPLER', '')))
            ws1.cell(row=row, column=7, value=str(result.get('YAPTIRIMLI_GTIPLER', '')))
            ws1.cell(row=row, column=8, value=str(result.get('GÜVEN_SEVİYESİ', '')))
            ws1.cell(row=row, column=9, value=str(result.get('AI_AÇIKLAMA', '')))
            ws1.cell(row=row, column=10, value=str(result.get('AI_TAVSIYE', '')))
            ws1.cell(row=row, column=11, value=str(result.get('BAŞLIK', '')))
            ws1.cell(row=row, column=12, value=str(result.get('URL', '')))
        
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        print(f"✅ Excel raporu oluşturuldu: {filename}")
        return filename
        
    except Exception as e:
        print(f"❌ Excel rapor hatası: {e}")
        return None

def display_results(results, company, country):
    """Sonuçları göster"""
    print(f"\n{'='*80}")
    print(f"📊 GERÇEK VERİTABANI ANALİZ SONUÇLARI: '{company}' ↔ {country}")
    print(f"{'='*80}")
    
    if not results:
        print("❌ Gerçek ticaret verisi bulunamadı!")
        return
    
    total_results = len(results)
    high_risk_count = len([r for r in results if r.get('YAPTIRIM_RISKI') == 'YÜKSEK'])
    medium_risk_count = len([r for r in results if r.get('YAPTIRIM_RISKI') == 'ORTA'])
    country_connection_count = len([r for r in results if r.get('ULKE_BAGLANTISI') == 'EVET'])
    
    print(f"\n📈 ÖZET:")
    print(f"   • Toplam Gerçek Veri: {total_results}")
    print(f"   • Ülke Bağlantısı: {country_connection_count}")
    print(f"   • YÜKSEK Yaptırım Riski: {high_risk_count}")
    print(f"   • ORTA Risk: {medium_risk_count}")
    
    if high_risk_count > 0:
        print(f"\n⚠️  KRİTİK RİSK UYARISI:")
        for result in results:
            if result.get('YAPTIRIM_RISKI') == 'YÜKSEK':
                print(f"   🔴 {result.get('BAŞLIK', '')[:60]}...")
                print(f"      Risk: {result.get('AI_AÇIKLAMA', '')}")
    
    for i, result in enumerate(results, 1):
        print(f"\n🔍 SONUÇ {i}:")
        print(f"   📝 Başlık: {result.get('BAŞLIK', 'N/A')}")
        print(f"   🌐 URL: {result.get('URL', 'N/A')}")
        print(f"   🎯 Durum: {result.get('DURUM', 'N/A')}")
        print(f"   ⚠️  Yaptırım Riski: {result.get('YAPTIRIM_RISKI', 'N/A')}")
        print(f"   🔗 Ülke Bağlantısı: {result.get('ULKE_BAGLANTISI', 'N/A')}")
        print(f"   🔍 GTIP Kodları: {result.get('TESPIT_EDILEN_GTIPLER', 'Yok')}")
        if result.get('YAPTIRIMLI_GTIPLER'):
            print(f"   🚫 Yasaklı GTIP: {result.get('YAPTIRIMLI_GTIPLER', '')}")
        print(f"   📊 Güven Seviyesi: {result.get('GÜVEN_SEVİYESİ', 'N/A')}")
        print(f"   📍 Kaynak Tipi: {result.get('KAYNAK_TİPİ', 'N/A')}")
        print(f"   💡 Açıklama: {result.get('AI_AÇIKLAMA', 'N/A')}")
        print(f"   💭 Tavsiye: {result.get('AI_TAVSIYE', 'N/A')}")
        print(f"   {'─'*60}")

def main():
    print("📊 GERÇEK TİCARET VERİTABANI ANALİZ SİSTEMİ")
    print("🎯 HEDEF: Gerçek ticaret veritabanlarına direkt bağlantı")
    print("💡 AVANTAJ: Eximpedia, Trademo, ExportGenius, Kompass")
    print("🔍 Kaynak: Gerçek ticaret verileri\n")
    
    config = Config()
    analyzer = SmartTradeAnalyzer(config)
    
    company = input("Şirket adını girin (TAM İSİM): ").strip()
    country = input("Ülke adını girin: ").strip()
    
    if not company or not country:
        print("❌ Şirket ve ülke bilgisi gereklidir!")
        return
    
    print(f"\n🚀 GERÇEK VERİTABANI ANALİZİ BAŞLATILIYOR: '{company}' ↔ {country}")
    print("⏳ Gerçek ticaret veritabanları taranıyor...")
    print("   Eximpedia, Trademo, ExportGenius, Kompass...\n")
    
    start_time = time.time()
    results = analyzer.smart_analyze(company, country)
    execution_time = time.time() - start_time
    
    if results:
        display_results(results, company, country)
        
        filename = create_excel_report(results, company, country)
        
        if filename:
            print(f"\n✅ Excel raporu oluşturuldu: {filename}")
            print(f"⏱️  Toplam çalışma süresi: {execution_time:.2f} saniye")
            
            try:
                open_excel = input("\n📂 Excel dosyasını açmak ister misiniz? (e/h): ").strip().lower()
                if open_excel == 'e':
                    if os.name == 'nt':
                        os.system(f'start excel "{filename}"')
                    elif os.name == 'posix':
                        os.system(f'open "{filename}"' if sys.platform == 'darwin' else f'xdg-open "{filename}"')
                    print("📂 Excel dosyası açılıyor...")
            except Exception as e:
                print(f"⚠️  Dosya otomatik açılamadı: {e}")
                print(f"📁 Manuel açın: {filename}")
        else:
            print("❌ Excel raporu oluşturulamadı!")
    else:
        print("❌ Gerçek ticaret verisi bulunamadı!")

if __name__ == "__main__":
    main()
